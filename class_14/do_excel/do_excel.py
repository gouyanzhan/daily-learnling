from openpyxl import  load_workbook
from class_14.do_excel.read_config import ReadConfig
class  DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self,button = 'all'):
        '''button:控制是否执行所有的用例，默认值为all，就执行所有用例
           如果不等于all，就进入分值判断
        button的值  只能输入all  列表 这两种类型的参数'''

        # 从配置文件读取button值
        button = ReadConfig().read_config('case.config','Mode','mode')
        wb = load_workbook(self.file_name)

        sheet = wb[self.sheet_name]  #不要写成字符串
        # res = sheet.cell(1,1).value
        #怎么拿到第一行的所有数据
        test_data = []
        for i in range(2,sheet.max_row + 1):  #(1,5) 1 2 3 4
            sub_data = {}
            sub_data['case_id'] = sheet.cell(i,1).value
            sub_data['module'] = sheet.cell(i, 2).value
            sub_data['title'] = sheet.cell(i, 3).value
            method = sheet.cell(i,4).value
            sub_data['method'] = method
            url = sheet.cell(i,5).value
            sub_data['url'] = url
            data = sheet.cell(i,6).value
            sub_data['data'] = data
            expected = sheet.cell(i,7).value
            sub_data['expected'] = expected
            test_data.append(sub_data)  #存储了所有数据
        #根据button值进行判断
        if button =='all':  #执行所有用例
            final_data = test_data
        else: #[1,2,3,4]
            final_data = []
            for item in test_data:  #对test_data  所有的测试数据 进行遍历
                # if item['case_id'] in button:
                if item['case_id'] in eval(button):   #因为从配置文件获取的是一个str值，需要转下类型
                    final_data.append(item)

        return final_data #返回获取到的数据

if __name__ == '__main__':
    DoExcel("test3.xlsx","python").get_data()