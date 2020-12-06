#地址  测试数据  断言期望结果   除了这几个不同  其他的都是高度相似的
#参数化 url  data  expected

#存到Excel 里面  python去操作Excel
#1：.xlsx  openpyxl 只支持这种格式
#2：创建

from openpyxl import  load_workbook
wb = load_workbook("test.xlsx")  #open the  given  filename  and  return the workbook
# file = open("test.xlsx")

#2：定位表单
sheet= wb["gyz"] #传表单名  返回一个表单对象

#3:定位单元格  行列值
res = sheet.cell(1,1).value

print("最大行:{0}".format(sheet.max_row))  #求表单的最大行
print("最大列:{0}".format(sheet.max_column)) #求表单的最大列
# print("拿到的结果是：",res)


#数据从Excel里面拿出来是什么类型
#数字还是数字  其他的都是字符串
print("url:{0},类型是{1}".format((sheet.cell(1,1).value),type(sheet.cell(1,1).value)))
print("url:{0},类型是{1}".format((sheet.cell(1,2).value),type(sheet.cell(1,2).value)))
print("url:{0},类型是{1}".format((sheet.cell(1,3).value),type(sheet.cell(1,3).value)))
print("url:{0},类型是{1}".format((sheet.cell(1,4).value),type(sheet.cell(1,4).value)))

#homework
#数据参数化，存到excel里，利用openpyxl写一个专门读取excel里面测试数据的类
# 通过初始化传参的方法，完成单元测试
#提交操作excel的类，test_suite,test_http类（类名可以自己定）
