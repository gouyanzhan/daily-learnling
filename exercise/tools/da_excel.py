
from openpyxl import  load_workbook

# 相对路径  绝对路径
class DoExcel:
    @staticmethod
    def get_data(self,file_name,sheet_name):
        # wb = load_workbook("../test_data/test_data.xlsx")  #linux路径-相对路径
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]

        test_data=[]
        for i in range(2,sheet.max_row+1):
            row_data = {}
            row_data['case_id'] = sheet.cell(i,1).value
            row_data['url'] = sheet.cell(i,2).value
            row_data['data'] = sheet.cell(i,3).value
            row_data['title'] = sheet.cell(i,4).value
            row_data['http_method'] = sheet.cell(i,5).value
            row_data['expected'] = sheet.cell(i,6).value # 添加一个期望值到测试数据中
            test_data.append(row_data)
        return test_data
    def write_back(self,file_name,sheet_name,i,value):  #专门写回数据
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i,7).value = value
        wb.save(file_name) #保存结果

if __name__ == '__main__':
    test_data = DoExcel().get_data("../test_data/test_data.xlsx","login")
    print(test_data)