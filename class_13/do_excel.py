from openpyxl import  load_workbook


class  DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        wb = load_workbook(self.file_name)

        sheet = wb[self.sheet_name]  #不要写成字符串
        # res = sheet.cell(1,1).value
        #怎么拿到第一行的所有数据
        test_data = []
        for i in range(1,sheet.max_row + 1):  #(1,5) 1 2 3 4
            sub_data = {}
            method = sheet.cell(i,1).value
            sub_data['method'] = method
            url = sheet.cell(i,2).value
            sub_data['url'] = url
            data = sheet.cell(i,3).value
            sub_data['data'] = data
            expected = sheet.cell(i,4).value
            sub_data['expected'] = expected
            test_data.append(sub_data)
        return(test_data)  #返回获取到的数据

if __name__ == '__main__':
    DoExcel("test.xlsx","python").get_data()