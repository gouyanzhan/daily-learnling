from openpyxl import  load_workbook

#读取数据的 法一 法二 都ok  以后肯定是法一
#法一 一次性读取所有数据 对内存的要求会高点
#法二 需要用的时候读取所有的数据  就是磁盘读写要求高点
# 磁盘-->内存-->cpu

#法三：do_excel3

class  DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.sheet_obj = load_workbook(self.file_name)[self.sheet_name]  #获取一个表单对象
        self.max_row = self.sheet_obj.max_row  #根据表单获取最大行数

    def get_data(self,i,j):
        #根据传入的坐标来获取值
        return self.sheet_obj.cell(i,j).value
    # def get_max_row(self):
    #     #根据表单获取最大行数
    #     return self.sheet_obj.max_row

if __name__ == '__main__':
    res =  DoExcel("test.xlsx","python").get_data(1,1)
    print(res)

