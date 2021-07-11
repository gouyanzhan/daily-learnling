from openpyxl import  load_workbook
#法三 仅供参考，日后优化

class  DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
    def get_header(self):
        #获取第一行的标题行
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]  # 不要写成字符串
        header = [] #存储标题行
        for j in range(1,sheet.max_column+1):
            header.append(1,j)
        return header

    def get_data(self):
        # 根据嵌套循环获取数据
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]  #不要写成字符串

        header = self.get_header()  #拿到header  是一个列表
        test_data = []

        for i in range(2,sheet.max_row + 1):  #(1,5) 1 2 3 4
            sub_data = {}
            for j in range(1,sheet.max_column + 1):
                # i = 1 j = 1 2 3 4 5 6 7
                sub_data[header[j-1]] = sheet.cell(i,j).value
            test_data.append(sub_data)

        return(test_data)  #返回获取到的数据

if __name__ == '__main__':
    DoExcel("test.xlsx","python").get_data()